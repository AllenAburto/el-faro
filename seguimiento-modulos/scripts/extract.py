#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extract.py — Genera js/app-data.js a partir del Excel maestro.

Plan de Trabajo — Fase 2 (P2-02): formaliza como script versionado el
proceso que hasta ahora era manual (abrir el Excel, copiar/recalcular y
pegar datos a mano en app-data.js). A partir de ahora, `app-data.js` se
considera un *artefacto generado* — el archivo fuente de verdad es
`Seguimiento_Nuevos_Modulos_V2_Dashboard.xlsx`.

Uso:
    python3 scripts/extract.py
    python3 scripts/extract.py --xlsx otra_copia.xlsx --out js/app-data.js

El GitHub Action de despliegue (`.github/workflows/deploy-pages.yml`)
ejecuta este script antes de publicar, así el sitio publicado siempre
refleja el Excel versionado en el repo — no una copia manual que se
puede desincronizar. `validate.py` corre justo después, como control de
calidad: si detecta datos inconsistentes, el build falla y no se publica
nada roto.
"""
import argparse
import datetime
import json
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
ROOT_DIR = SCRIPT_DIR.parent  # seguimiento-modulos/

HEADER_COMMENT = """/* ======================================================== app-data.js
   Datos extraídos de Seguimiento_Nuevos_Modulos_V2_Dashboard.xlsx
   (generado automáticamente por scripts/extract.py — no editar a mano)
   ========================================================= */
"""


def cell(v):
    """Normaliza un valor de celda: las fechas se serializan como ISO (YYYY-MM-DD)."""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return v


def find_title_row(ws, title, col=1):
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=col).value == title:
            return r
    raise ValueError(f"No se encontró el título {title!r} en la hoja {ws.title!r}")


def read_table_by_title(ws, title, title_col=1, ncols=None):
    """título en fila T -> encabezados en fila T+2 -> datos desde T+3 hasta la primera fila en blanco."""
    t = find_title_row(ws, title, col=title_col)
    hdr_row = t + 2
    headers = []
    c = 1
    while True:
        v = ws.cell(row=hdr_row, column=c).value
        if v is None:
            if ncols and c <= ncols:
                headers.append(None)
                c += 1
                continue
            break
        headers.append(v)
        c += 1
    rows = []
    r = hdr_row + 1
    while True:
        first = ws.cell(row=r, column=1).value
        if first is None:
            break
        rowvals = [ws.cell(row=r, column=i + 1).value for i in range(len(headers))]
        rows.append({headers[i]: cell(rowvals[i]) for i in range(len(headers)) if headers[i]})
        r += 1
    return rows


def read_kv_by_title(ws, title, title_col=1):
    t = find_title_row(ws, title, col=title_col)
    r = t + 1
    out = {}
    blanks = 0
    while blanks < 2:
        k = ws.cell(row=r, column=1).value
        if k is None:
            blanks += 1
            r += 1
            continue
        blanks = 0
        out[k] = cell(ws.cell(row=r, column=2).value)
        r += 1
    return out


def read_req_sheet(wb, name):
    ws = wb[name]
    max_col = ws.max_column
    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        rowvals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if all(v is None for v in rowvals):
            continue
        rows.append({headers[i]: cell(rowvals[i]) for i in range(len(headers)) if headers[i]})
    return rows


def read_mini_list(ws, col, row_start=5):
    items = []
    r = row_start
    while True:
        v = ws.cell(row=r, column=col).value
        if v is None:
            break
        items.append(v.lstrip("•").strip())
        r += 1
    return items


def find_row_with_value(ws, value, col):
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=col).value == value:
            return r
    return None


def extract(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = {}

    # -------------------------------------------------------- Indicadores
    wi = wb["Indicadores"]
    ejec = read_kv_by_title(wi, "Resumen Ejecutivo")
    data["dashboard_meta"] = {
        "total": ejec.get("Total actividades"),
        "finalizados": ejec.get("Finalizadas"),
        "finalizados_desfase": ejec.get("Finalizadas con desfase"),
        "en_curso": ejec.get("En curso"),
        "inicio_retrasado": ejec.get("Inicio retrasado (en curso)"),
        "programado": ejec.get("Programadas"),
        "no_iniciado": ejec.get("No iniciadas"),
        "atrasados": ejec.get("Atrasadas"),
        "no_planificado": ejec.get("No planificadas"),
        "completar_info": ejec.get("COMPLETAR (falta info)"),
        "avance_progreso": ejec.get("% avance"),
        "estado_general": ejec.get("Estado general"),
    }
    data["resumen_modulo"] = read_table_by_title(wi, "Resumen por Módulo")
    data["resumen_responsable"] = read_table_by_title(wi, "Resumen por Responsable")
    data["resumen_etapa"] = read_table_by_title(wi, "Resumen por Etapa")
    data["resumen_componente"] = [
        {k: v for k, v in row.items() if k != "Barra"}
        for row in read_table_by_title(wi, "Resumen por Componente / Subetapa")
    ]
    data["req_por_modulo"] = read_table_by_title(wi, "Resumen de Requerimientos por Módulo")
    data["avance_global_requerimientos"] = read_kv_by_title(wi, "Avance Global de Requerimientos")

    # Nota: "estado_general" y "plazos" quedan también congelados aquí (tal
    # como los calcula el Excel), pero el portal YA NO los usa para el
    # semáforo del hero — desde el Plan de Trabajo Fase 0 (P0-01),
    # UI.computeSemaforo() en common.js los recalcula en cada carga contra
    # la Bitácora real. Se conservan en el JSON solo como referencia/histórico.
    atrasadas = ejec.get("Atrasadas") or 0
    if atrasadas == 0:
        plazos = "Verde"
    elif atrasadas <= 2:
        plazos = "Amarillo"
    else:
        plazos = "Rojo"
    data["dashboard_meta"]["plazos"] = plazos
    data["dashboard_meta"]["actualizado"] = datetime.date.today().isoformat()

    # ------------------------------------------------------------ Etapas
    we = wb["Etapas"]
    headers = [we.cell(row=1, column=c).value for c in range(1, 18)]
    etapas = []
    for r in range(2, we.max_row + 1):
        if we.cell(row=r, column=2).value is None and we.cell(row=r, column=4).value is None:
            continue
        rowvals = [we.cell(row=r, column=c).value for c in range(1, 18)]
        rec = {headers[i]: cell(rowvals[i]) for i in range(len(headers))}
        # La columna "N°" viene vacía en el Excel de origen (no es un id
        # útil); se genera un id estable a partir de la fila, usado por el
        # sistema de edición local (Store) del portal.
        rec["id"] = f"et-{r}"
        etapas.append(rec)
    data["etapas"] = etapas

    # ----------------------------------------------------------- Req. *
    data["req_autoatencion"] = read_req_sheet(wb, "Req. AutoAtención")
    data["req_rad"] = read_req_sheet(wb, "Req. RAD")
    data["req_lm"] = read_req_sheet(wb, "Req. LM")
    data["req_calificaciones"] = read_req_sheet(wb, "Req. Calificaciones")
    data["req_forcap"] = read_req_sheet(wb, "Req. FORCAP")
    data["req_portal"] = read_req_sheet(wb, "Req. Portal")
    data["req_vals"] = read_req_sheet(wb, "Req. VALS")

    # --------------------------------------------------------- Bitácora
    wb_bit = wb["Bitácora"]
    bit_keys = [
        "n", "fecha_registro", "modulo", "etapa", "componente", "origen", "tipo_registro",
        "descripcion", "impacto", "productos", "codigo_link", "responsable", "accion",
        "fecha_comprometida", "fecha_real", "fecha_real_cierre", "estado", "observaciones",
        "dias_retraso",
    ]
    bitacora = []
    for r in range(2, wb_bit.max_row + 1):
        rowvals = [wb_bit.cell(row=r, column=c).value for c in range(1, len(bit_keys) + 1)]
        if all(v is None for v in rowvals):
            continue
        rec = {k: cell(v) for k, v in zip(bit_keys, rowvals)}
        rec["id"] = f"bit-{rec['n']}"
        bitacora.append(rec)
    data["bitacora"] = bitacora

    # ---------------------------------------------------- Tablas y Glosario
    wt = wb["Tablas y Glosario"]
    data["tipos"] = [{"Tipo": t} for t in read_mini_list(wt, 2)]
    data["modulos_lista"] = [{"Módulo": t} for t in read_mini_list(wt, 4)]
    data["etapas_lista"] = [{"Etapa": t} for t in read_mini_list(wt, 6)]

    g_title = find_row_with_value(wt, "Glosario", 2)
    g_hdr = g_title + 1
    r = g_hdr + 1
    glosario = []
    while True:
        term = wt.cell(row=r, column=2).value
        if term is None:
            break
        defin = wt.cell(row=r, column=3).value
        glosario.append({"Termino": term, "Definicion": defin})
        r += 1
    data["glosario"] = glosario

    raci_title_row = None
    for r in range(1, wt.max_row + 1):
        v = wt.cell(row=r, column=2).value
        if isinstance(v, str) and v.startswith("Matriz RACI —"):
            raci_title_row = r
            break
    raci_hdr = raci_title_row + 1
    raci_cols = [2, 7, 8, 9, 10, 11]
    raci_headers = [wt.cell(row=raci_hdr, column=c).value for c in raci_cols]
    raci = []
    r = raci_hdr + 1
    while True:
        act = wt.cell(row=r, column=2).value
        if act is None:
            break
        row = {raci_headers[0]: act}
        for i, c in enumerate(raci_cols[1:], start=1):
            row[raci_headers[i]] = wt.cell(row=r, column=c).value
        raci.append(row)
        r += 1
    data["raci_ejemplo"] = raci

    return data


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--xlsx", default=str(ROOT_DIR / "Seguimiento_Nuevos_Modulos_V2_Dashboard.xlsx"),
        help="Ruta al Excel maestro (por defecto: el archivo versionado en el repo).",
    )
    ap.add_argument(
        "--out", default=str(ROOT_DIR / "js" / "app-data.js"),
        help="Ruta de salida para app-data.js (por defecto: js/app-data.js).",
    )
    ap.add_argument(
        "--json-out", default=None,
        help="Si se indica, además escribe los datos crudos como JSON en esta ruta (útil para validate.py o debug).",
    )
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"No se encontró el Excel maestro: {xlsx_path}")

    data = extract(xlsx_path)

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    body = json.dumps(data, ensure_ascii=False)
    out_path.write_text(HEADER_COMMENT + "window.APP_DATA = " + body + ";\n", encoding="utf-8")

    if args.json_out:
        json_path = Path(args.json_out)
        json_path.parent.mkdir(parents=True, exist_ok=True)
        json_path.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"OK — {out_path} generado desde {xlsx_path.name}")
    print(f"  etapas: {len(data['etapas'])}")
    print(f"  bitacora: {len(data['bitacora'])}")
    for k in ("req_autoatencion", "req_rad", "req_lm", "req_calificaciones", "req_forcap", "req_portal", "req_vals"):
        print(f"  {k}: {len(data[k])}")
    print(f"  glosario: {len(data['glosario'])} términos")
    print(f"  dashboard_meta: {data['dashboard_meta']}")


if __name__ == "__main__":
    main()
